import os
import numpy as np
import csv

# 新增：写 xlsx 用
from openpyxl import Workbook
from openpyxl.styles import Font

# ===================== 配置区：这里自由切换 =====================
# 数据集名称（你原来是 XJTU）
DATA = "XJTU"

# 输入类型：在下面三种里切换即可
#   "charge"
#   "partial_charge"
#   "handcraf_features"
INPUT_TYPE = "partial_charge"  # <<< 改成你想要的：partial_charge / handcraf_features

# 基础路径：对应你训练脚本里的 save_folder = results/{data}-{input_type}/...
BASE_DIR = os.path.join("results", f"{DATA}-{INPUT_TYPE}")

# 结果文件名模板：自动适配 charge / partial_charge / handcraf_features
#   例如：LSTM_charge_results.npz, LSTM_partial_charge_results.npz ...
RESULT_FILE_TEMPLATE = "{model}_" + INPUT_TYPE + "_results.npz"

# 输出文件名前缀
OUT_PREFIX = f"{DATA}_{INPUT_TYPE}_table"
# ============================================================


# 5 个模型，对应 Table C.10 的列顺序
MODELS = ["CNN", "LSTM", "GRU", "MLP", "Attention"]

# XJTU 各个 batch 的电池数
BATCH_TO_IDS = {
    1: range(1, 9),
    2: range(1, 16),
    3: range(1, 9),
    4: range(1, 9),
    5: range(1, 9),
    6: range(1, 9),
}

N_EXP = 3  # 每个 (batch, battery, model) 的实验次数


def load_errors_for_cell(model, batch, battery, n_exp=N_EXP):
    all_errs = []

    for e in range(1, n_exp + 1):
        # 根据模板自动拼 npz 文件名
        result_fname = RESULT_FILE_TEMPLATE.format(model=model)

        npz_path = os.path.join(
            BASE_DIR,
            model,
            f"batch{batch}-testbattery{battery}",
            f"experiment{e}",
            result_fname,
        )

        if not os.path.exists(npz_path):
            print(f"[WARN] Missing file: {npz_path}")
            continue

        data = np.load(npz_path)
        if "test_errors" not in data:
            print(f"[WARN] 'test_errors' not in {npz_path}")
            continue

        err = data["test_errors"][:3]  # MAE, MAPE, MSE
        all_errs.append(err)

    if len(all_errs) == 0:
        return np.array([np.nan, np.nan, np.nan], dtype=float)

    all_errs = np.stack(all_errs, axis=0)
    mean_err = all_errs.mean(axis=0)
    return mean_err * 1000.0  # 放大 1000


def safe_sheet_title(data, input_type):
    # 原始标题
    raw = f"{data}_{input_type}_Table_C10_like"
    # Excel 限制：<=31
    if len(raw) <= 31:
        return raw

    # 太长就缩短：保留前缀 + 最关键的 C10
    short = f"{data}_{input_type}_C10"
    if len(short) <= 31:
        return short

    # 还超就粗暴截断
    return short[:31]

def main():
    out_csv = f"{OUT_PREFIX}.csv"
    out_xlsx = f"{OUT_PREFIX}.xlsx"

    header = ["Batch", "Battery"]
    for m in MODELS:
        header.extend([f"{m}_MAE", f"{m}_MAPE", f"{m}_MSE"])

    rows = []

    for batch, ids_list in BATCH_TO_IDS.items():
        for battery in ids_list:
            row = [batch, battery]

            for model in MODELS:
                mae, mape, mse = load_errors_for_cell(model, batch, battery)
                row.extend([
                    float(f"{mae:.3f}") if not np.isnan(mae) else np.nan,
                    float(f"{mape:.3f}") if not np.isnan(mape) else np.nan,
                    float(f"{mse:.3f}") if not np.isnan(mse) else np.nan,
                ])

            rows.append(row)

    # ========= 保存 CSV =========
    os.makedirs(os.path.dirname(out_csv) or ".", exist_ok=True)
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header)

        for row in rows:
            csv_row = [("" if (isinstance(v, float) and np.isnan(v)) else v) for v in row]
            writer.writerow(csv_row)

    # ========= 保存 XLSX，并加粗最优值 =========
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(DATA, INPUT_TYPE)

    ws.append(header)
    bold_font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        ws.append([None if (isinstance(v, float) and np.isnan(v)) else v for v in row])

        # 分 metric 取最优
        metrics_vals = {0: [], 1: [], 2: []}
        metrics_cols = {0: [], 1: [], 2: []}

        for m_i, _ in enumerate(MODELS):
            base_col = 3 + m_i * 3
            for idx in range(3):
                cell_val = ws.cell(row=row_idx, column=base_col + idx).value
                metrics_vals[idx].append(cell_val)
                metrics_cols[idx].append(base_col + idx)

        # 加粗最优值
        for k in range(3):
            vals_cols = [(v, c) for v, c in zip(metrics_vals[k], metrics_cols[k])
                         if v is not None and not (isinstance(v, float) and np.isnan(v))]
            if not vals_cols:
                continue

            best_val = min(v for v, _ in vals_cols)

            for v, col in vals_cols:
                if abs(v - best_val) < 1e-6:
                    ws.cell(row=row_idx, column=col).font = bold_font

    wb.save(out_xlsx)

    print(f"DATA = {DATA}, INPUT_TYPE = {INPUT_TYPE}")
    print(f"Saved CSV → {out_csv}")
    print(f"Saved XLSX → {out_xlsx}")


if __name__ == "__main__":
    main()