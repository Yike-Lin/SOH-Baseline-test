import os
import numpy as np
import csv

# 新增：写 xlsx 用
from openpyxl import Workbook
from openpyxl.styles import Font

# 基础路径：根据你训练脚本里的 save_folder 来改
# 你之前用的是：results/{args.data}-{args.input_type}/...
# 对应 XJTU + charge = "results/XJTU-charge"
BASE_DIR = "results/XJTU-charge"

# 5 个模型，对应 Table C.10 的列顺序
MODELS = ["CNN", "LSTM", "GRU", "MLP", "Attention"]

# XJTU 各个 batch 的电池数
BATCH_TO_IDS = {
    1: range(1, 9),
    2: range(1, 16),
    3: range(1, 9),
    4: range(1, 9),
    5: range(1, 9),
    6: range(1, 9),
}

N_EXP = 3  # 每个 (batch, battery, model) 的实验次数，对应你训练里的 for e in range(3)

def load_errors_for_cell(model, batch, battery, n_exp=N_EXP):
    """
    读取某个 (model, batch, battery) 的多次实验结果，
    返回 (MAE, MAPE, MSE) 的平均值，并且 *1000（和表里展示一致）

    假设 npz 结构中:
      test_errors = [MAE, MAPE, MSE, R2] 或至少前三个是 MAE,MAPE,MSE
    """
    all_errs = []

    for e in range(1, n_exp + 1):
        # 对应你训练时的 save_folder
        npz_path = os.path.join(
            BASE_DIR,
            model,
            f"batch{batch}-testbattery{battery}",
            f"experiment{e}",
            f"{model}_charge_results.npz"
        )

        if not os.path.exists(npz_path):
            print(f"[WARN] Missing file: {npz_path}")
            continue

        data = np.load(npz_path)
        if "test_errors" not in data:
            print(f"[WARN] 'test_errors' not in {npz_path}")
            continue

        err = data["test_errors"]
        # 取前三个: MAE, MAPE, MSE
        err = err[:3]
        all_errs.append(err)

    if len(all_errs) == 0:
        # 没有任何有效实验，返回 NaN
        return np.array([np.nan, np.nan, np.nan], dtype=float)

    all_errs = np.stack(all_errs, axis=0)  # (n_valid_exp, 3)
    mean_err = all_errs.mean(axis=0)       # (3,)

    # 展示用放大 1000 倍（和 Appendix Table Note 一致）
    return mean_err * 1000.0


def main():
    # 输出文件名
    out_csv = "XJTU_table_C10_like.csv"
    out_xlsx = "XJTU_table_C10_like.xlsx"

    # 构造表头：Batch, Battery, 然后每个模型 3 列
    header = ["Batch", "Battery"]
    for m in MODELS:
        header.extend([
            f"{m}_MAE",
            f"{m}_MAPE",
            f"{m}_MSE",
        ])

    rows = []

    # 先把所有数据算出来，rows 是纯数值，不做格式
    for batch, ids_list in BATCH_TO_IDS.items():
        for battery in ids_list:
            row = [batch, battery]

            for model in MODELS:
                mae, mape, mse = load_errors_for_cell(model, batch, battery)
                # 保留 3 位小数，方便和论文里的数对比
                row.extend([
                    float(f"{mae:.3f}") if not np.isnan(mae) else np.nan,
                    float(f"{mape:.3f}") if not np.isnan(mape) else np.nan,
                    float(f"{mse:.3f}") if not np.isnan(mse) else np.nan,
                ])

            rows.append(row)

    # ---------- 1) 先照旧保存 CSV ----------
    os.makedirs(os.path.dirname(out_csv) or ".", exist_ok=True)
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        # CSV 表头里的模型列名用字符串
        csv_header = ["Batch", "Battery"]
        for m in MODELS:
            csv_header.extend([
                f"{m}_MAE",
                f"{m}_MAPE",
                f"{m}_MSE",
            ])
        writer.writerow(csv_header)

        # 写数据行（转换 NaN 为 ""）
        for row in rows:
            csv_row = []
            for v in row:
                if isinstance(v, float) and np.isnan(v):
                    csv_row.append("")
                else:
                    csv_row.append(v)
            writer.writerow(csv_row)

    # ---------- 2) 再保存为 XLSX，并对每行的最优值加粗 ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "XJTU_Table_C10_like"

    # 写表头
    ws.append(csv_header)

    # openpyxl 的 Font 对象，用于加粗
    bold_font = Font(bold=True)

    # 从第二行开始写数据（第一行是 header）
    for row_idx, row in enumerate(rows, start=2):
        ws_row_vals = []

        # 先把这一整行写进去（不加粗）
        for v in row:
            if isinstance(v, float) and np.isnan(v):
                ws_row_vals.append(None)
            else:
                ws_row_vals.append(v)
        ws.append(ws_row_vals)

        # 现在对这一行做“最小值加粗”
        # 列结构：
        # col 1: Batch
        # col 2: Battery
        # col 3~5: CNN_MAE/MAPE/MSE
        # col 6~8: LSTM_...
        # ...
        # 一共 2 + 5*3 = 17 列
        # 对同一行里：
        #   - 找 5 个模型在 "MAE" 列的最小值 → 加粗
        #   - 找 5 个模型在 "MAPE" 列的最小值 → 加粗
        #   - 找 5 个模型在 "MSE" 列的最小值 → 加粗

        # 先把每个 metric 的值按模型收集起来
        # metrics_vals[metric_index][model_index] = value
        # metric_index: 0 -> MAE, 1 -> MAPE, 2 -> MSE
        metrics_vals = {0: [], 1: [], 2: []}
        metrics_cols = {0: [], 1: [], 2: []}  # 记录每个值所在列号

        for m_i, _ in enumerate(MODELS):
            start_col = 3 + m_i * 3  # 这一组模型的 MAE 列
            # MAE
            mae_cell = ws.cell(row=row_idx, column=start_col)
            metrics_vals[0].append(mae_cell.value)
            metrics_cols[0].append(start_col)
            # MAPE
            mape_cell = ws.cell(row=row_idx, column=start_col + 1)
            metrics_vals[1].append(mape_cell.value)
            metrics_cols[1].append(start_col + 1)
            # MSE
            mse_cell = ws.cell(row=row_idx, column=start_col + 2)
            metrics_vals[2].append(mse_cell.value)
            metrics_cols[2].append(start_col + 2)

        # 对每个 metric（MAE/MAPE/MSE）分别找最小值并加粗
        for metric_idx in [0, 1, 2]:
            vals = metrics_vals[metric_idx]
            cols = metrics_cols[metric_idx]

            # 去掉 None / NaN
            valid_pairs = [
                (v, c) for v, c in zip(vals, cols)
                if v is not None and not (isinstance(v, float) and np.isnan(v))
            ]
            if not valid_pairs:
                continue

            best_val = min(v for v, _ in valid_pairs)

            # 允许一点浮点误差，谁 == best 就加粗
            for v, c in valid_pairs:
                if abs(v - best_val) < 1e-6:
                    cell = ws.cell(row=row_idx, column=c)
                    cell.font = bold_font

    # 保存 xlsx
    wb.save(out_xlsx)

    print(f"Done. Saved CSV to:   {out_csv}")
    print(f"Done. Saved XLSX to:  {out_xlsx}")
    print("XLSX 里每行的每个 metric（MAE/MAPE/MSE），5 个模型中最小值已用粗体标出。")


if __name__ == "__main__":
    main()